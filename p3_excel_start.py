# P3 EXCEL
# Brinley Gregory, Ethan Carn, Luke Miller
# Madi Diefenbach, Seth Mortensen, Sydney Trojahn Hedges
# IS 303 Section 004
# PROJECT DESCRIPTION: 
# Takes poorly organized data from a excel workbook and organizes the data into a new workbook. 

# from instructions
import openpyxl 
from openpyxl import Workbook
from openpyxl.styles import Font

# imports student class from studentClass.py for use in main program file
from studentClass import Student

# gives user a choice of which file to use
print("1: Poorly_Organized_Data_1.xlsx")
print("2: Poorly_Organized_Data_2.xlsx")
testChoice = int(input("Choose which data you'd like to format (1 or 2): "))

if testChoice == 1:
    poorWorkbook = openpyxl.load_workbook("Poorly_Organized_Data_1.xlsx")
elif testChoice == 2:
    poorWorkbook = openpyxl.load_workbook("Poorly_Organized_Data_2.xlsx")

sheet = poorWorkbook.active


# step 2 in to-do Use a for loop to get every row of data in the poorly formatted excel worksheet.
def getStudentObjects (sheet) :
    iRow = 2 
    studentList = []
    # create while loop to run as long as the row has data
    while sheet["A" + str(iRow)].value != None :
        # gather variables
        subject = sheet["A" + str(iRow)].value 
        studentData = sheet["B" + str(iRow)].value
        grade = sheet["C" + str(iRow)].value
        # create student object passing it the parameters of subject, studentData, and grade
        student = Student(subject, studentData, grade)
        # add student object to list of students
        studentList.append(student)
        # move to the next row
        iRow += 1
    # return the student list
    return studentList


# THIS FUNCTION SHOULD RETURN A LIST OF STUDENT OBJECTS FROM THE SELECTED EXCEL SHEET
studentList = getStudentObjects(sheet)

# create function that creates the worksheets and names them based on the subject
def createWorksheets (studentList) :
    # create an workbook to store organized data
    organizedWorkbook = Workbook () 
    # for each student in the list of students 
    for student in range(len(studentList)) :
        # if the subject name is not a worksheet name enter 
        if studentList[student].class_name not in organizedWorkbook.sheetnames :
            # creates a worksheet titles the subject
            organizedWorkbook.create_sheet(title=studentList[student].class_name)
            print(f"{studentList[student].class_name} was added as a worksheet.")
        # deletes the original worksheet
        if "Sheet" in organizedWorkbook.sheetnames :
            del organizedWorkbook["Sheet"]
    # saves the workbook 
    organizedWorkbook.save(filename="Organized_Data.xlsx")
    organizedWorkbook.close ()

#call createworksheets function
createWorksheets(studentList)

# THIS FUNCTION SHOULD ADD ALL THE STUDENT DATA TO THE NEW FILE AND CORRECT CLASS SHEETS
def addStudentData(studentList, organizedWorkbook):
    # open workbook
    wb = openpyxl.load_workbook(organizedWorkbook)

    # create headers
    headers = ["Last Name", "First Name", "Student ID", "Grade"]

    # loop through every sheet in the workbook
    for sheet_name in wb.sheetnames:
        # activate worksheet
        worksheet = wb[sheet_name]
        wb.active = worksheet  

        # put headers on every sheet
        for col, header in enumerate(headers, start=1):
            worksheet.cell(row=1, column=col, value=header)
        
        # loop through every student object and see if they are in that class
        # if student is in the class, input their data into the sheet
        for student in range(len(studentList)):
            if wb.active.title == studentList[student].class_name:
                iRow = wb.active.max_row + 1  # Get the first empty row
                iCol = "A"
                wb.active[iCol + str(iRow)].value = studentList[student].lname
                iCol = "B"
                wb.active[iCol + str(iRow)].value = studentList[student].fname
                iCol = "C"
                wb.active[iCol + str(iRow)].value =studentList[student].student_id
                iCol = "D"
                wb.active[iCol + str(iRow)].value = studentList[student].grade

    # saves the workbook 
    wb.save(organizedWorkbook)
    wb.close()

# calles addstudentdata function
organizedWorkbook = "Organized_Data.xlsx"
addStudentData(studentList,organizedWorkbook)

# add filter function
def addFilter (OrganizedWorkbook) :
# load workbook
    wb = openpyxl.load_workbook(OrganizedWorkbook)
# for each worksheet in the workbook
    for worksheet in wb.worksheets :
        # make the current worksheet
        currWorksheet = worksheet
        # find the max row
        max_row = currWorksheet.max_row
        # autofilter the worksheet
        currWorksheet.auto_filter.ref = f"A1:D{max_row}"
        wb.save("Organized_Data.xlsx")
        wb.close()

#call addFilter function
addFilter(organizedWorkbook)

# function that finds min, max, etc. data of each class and prints it in F and G
def addSummaries(organizedWorkbook):
    # create list with headers
    headers = ['Summary Statistics', 'Highest Grade', 'Lowest Grade', 'Mean Grade', 'Median Grade', 'Number of Students']

    # open workbook
    wb = openpyxl.load_workbook(organizedWorkbook)

    # loop through every sheet in the workbook
    for sheet_name in wb.sheetnames:
        # activate worksheet
        worksheet = wb[sheet_name]
        wb.active = worksheet  

        # put headers on every sheet in column F
        for row, header in enumerate(headers, start=1):
            worksheet.cell(row=row, column=6, value=header)
        
        # calculate data and add it to column G
        iMaxRow = worksheet.max_row
        worksheet["G1"] = "Value"
        worksheet["G2"] = f"=MAX(D2:D{iMaxRow})"
        worksheet["G3"] = f"=MIN(D2:D{iMaxRow})"
        worksheet["G4"] = f"=AVERAGE(D2:D{iMaxRow})"
        worksheet["G5"] = f"=MEDIAN(D2:D{iMaxRow})"
        worksheet["G6"] = f"=COUNTA(D2:D{iMaxRow})"
    
    #save workbook
    wb.save(organizedWorkbook)

# call summaries function
addSummaries(organizedWorkbook)

#Format columns 
def format_columns(organizedWorkbook): 
    wb = openpyxl.load_workbook(organizedWorkbook)
    cols = ['A', 'B', 'C', 'D', 'F', 'G']
    bold_font = Font(bold=True)
    for sheet in wb.worksheets:
        for col_letter in cols:
            cell = sheet[f'{col_letter}1']

            cell.font = bold_font

            header_text = str(cell.value) if cell.value else ""
            new_width = len(header_text) + 5
            sheet.column_dimensions[col_letter].width = new_width
            wb.save ("Organized_Data.xlsx")

# calls formt columns function
format_columns(organizedWorkbook)

# saves and closes the workbook 
wb = openpyxl.load_workbook(organizedWorkbook)
wb.save(organizedWorkbook)
wb.close()